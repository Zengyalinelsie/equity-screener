"""【成员/universe】不调 Wind, 由源表 + codes csv 重建 monthly_membership。

成员标志取自 config/codes_a.csv / codes_hk.csv (已含 in_index_* / in_hsi 等列);
A 股权重取自源表 `Index Future Valuation Table(Wind)_clean.xlsx` 的 50W/300W/500W/1000W
sheet (D 列 i_weight)。HK 无权重源 → w_* 留 NULL。

替代了原来月频 wset 模板 (make_template_index_const / pull_index_const), 少一个 macOS
Excel 上易碎的环节。源表/codes 刷新后重跑本脚本即可。

    python scripts/import_membership.py
"""
from __future__ import annotations

import csv
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent.resolve()
DB = ROOT / "data" / "wind_history.db"
SOURCE = ROOT / "templates" / "Index Future Valuation Table(Wind)_clean.xlsx"
CODES_A = ROOT / "config" / "codes_a.csv"
CODES_HK = ROOT / "config" / "codes_hk.csv"

# 源表 W sheet → 权重列
WSHEET_TO_WCOL = {"50W": "w_50", "300W": "w_300", "500W": "w_500", "1000W": "w_1000"}
# codes_a.csv 标志列 → schema 标志列
A_FLAG_MAP = {
    "in_index_50": "in_50", "in_index_300": "in_300",
    "in_index_500": "in_500", "in_index_1000": "in_1000",
}
HK_FLAG_MAP = {"in_hsi": "in_hsi", "in_hscei": "in_hscei", "in_hstech": "in_hstech"}
ALL_FLAGS = ["in_50", "in_300", "in_500", "in_1000", "in_hsi", "in_hscei", "in_hstech"]
ALL_WS = ["w_50", "w_300", "w_500", "w_1000", "w_hsi", "w_hscei", "w_hstech"]


def _num(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _read_weights() -> dict[str, dict[str, float]]:
    """{wind_code: {w_50: .., w_300: ..}} 从源表 W sheets。"""
    out: dict[str, dict[str, float]] = {}
    if not SOURCE.exists():
        print(f"⚠️  源表不存在, A 股权重留空: {SOURCE.name}")
        return out
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    for sheet, wcol in WSHEET_TO_WCOL.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=6, values_only=True):  # 数据从第 6 行
            if not row or len(row) < 4:
                continue
            code, weight = row[1], _num(row[3])  # B=wind_code, D=i_weight
            if not code or weight is None:
                continue
            out.setdefault(str(code).strip(), {})[wcol] = weight
    wb.close()
    return out


def _read_codes(path: Path, market: str, flag_map: dict[str, str]) -> dict[str, dict]:
    """{wind_code: {market, in_*: 0/1}} 从 codes csv。"""
    out: dict[str, dict] = {}
    if not path.exists():
        print(f"⚠️  {path.name} 不存在, 跳过 {market}")
        return out
    with open(path, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            code = (r.get("wind_code") or "").strip()
            if not code:
                continue
            rec = {"market": market}
            for src_col, dst_col in flag_map.items():
                rec[dst_col] = 1 if str(r.get(src_col, "")).strip() in ("1", "1.0", "True") else 0
            out[code] = rec
    return out


def _as_of_date() -> str:
    """取源表 50W!B2 的快照日期, 失败则今天。"""
    if SOURCE.exists():
        try:
            wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
            raw = str(wb["50W"]["B2"].value or "").strip()
            wb.close()
            if len(raw) == 8 and raw.isdigit():
                return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
        except Exception:
            pass
    return datetime.now().strftime("%Y-%m-%d")


def main():
    weights = _read_weights()
    members = {}
    members.update(_read_codes(CODES_A, "A", A_FLAG_MAP))
    members.update(_read_codes(CODES_HK, "HK", HK_FLAG_MAP))
    as_of = _as_of_date()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DB, timeout=60)
    cur = conn.cursor()
    cur.execute("DELETE FROM monthly_membership WHERE as_of_date = ?", (as_of,))

    n = 0
    for code, rec in members.items():
        flags = {k: rec.get(k, 0) for k in ALL_FLAGS}
        ws = {k: None for k in ALL_WS}
        ws.update(weights.get(code, {}))
        cur.execute(
            "INSERT OR REPLACE INTO monthly_membership "
            "(as_of_date, wind_code, market, in_50, in_300, in_500, in_1000, "
            " in_hsi, in_hscei, in_hstech, w_50, w_300, w_500, w_1000, "
            " w_hsi, w_hscei, w_hstech, last_update) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (as_of, code, rec["market"],
             flags["in_50"], flags["in_300"], flags["in_500"], flags["in_1000"],
             flags["in_hsi"], flags["in_hscei"], flags["in_hstech"],
             ws["w_50"], ws["w_300"], ws["w_500"], ws["w_1000"],
             ws["w_hsi"], ws["w_hscei"], ws["w_hstech"], now),
        )
        n += 1

    cur.execute(
        "INSERT OR REPLACE INTO pipeline_freshness (dataset, last_run, status, rows, notes) "
        "VALUES ('monthly_membership', ?, 'ok', ?, ?)",
        (now, n, f"as_of={as_of} (源表 W sheets + codes csv, 无 Wind)"),
    )
    conn.commit()
    conn.close()
    a = sum(1 for r in members.values() if r["market"] == "A")
    hk = sum(1 for r in members.values() if r["market"] == "HK")
    print(f"✅ monthly_membership as_of={as_of}: {n} 行 (A={a} HK={hk}), 带权重 A 股 {len(weights)} 只")


if __name__ == "__main__":
    main()
